from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.utils import timezone
from academics.models import Subject, Enrollment
from users.models import User
from announcements.models import Announcement, AnnouncementCategory

class FacultyRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_faculty

class FacultyAttendanceSubjectListView(LoginRequiredMixin, FacultyRequiredMixin, ListView):
    model = Subject
    template_name = 'attendance/faculty_subject_list.html'
    context_object_name = 'subjects'

    def get_queryset(self):
        # Subjects taught by this faculty
        return Subject.objects.filter(faculty=self.request.user, is_active=True)

class MarkAttendanceView(LoginRequiredMixin, FacultyRequiredMixin, View):
    def get(self, request, subject_id):
        subject = get_object_or_404(Subject, id=subject_id, faculty=request.user)
        date = request.GET.get('date', timezone.now().date().isoformat())
        
        # Get students enrolled in the course and semester of the subject
        students = User.objects.filter(
            role='STUDENT',
            enrollments__course=subject.course,
            enrollments__current_semester=subject.semester,
            enrollments__is_active=True
        ).distinct()

        # Populate map using current attendance percentage from User model
        attendance_map = {student.id: student.attendance for student in students}
        return render(request, 'attendance/mark_attendance.html', {
            'subject': subject,
            'students': students,
            'date': date,
            'attendance_map': attendance_map
        })

    def post(self, request, subject_id):
        subject = get_object_or_404(Subject, id=subject_id, faculty=request.user)
        date = request.POST.get('date')
        
        # Get list of students again to validate
        students = User.objects.filter(
            role='STUDENT',
            enrollments__course=subject.course,
            enrollments__current_semester=subject.semester,
            enrollments__is_active=True
        ).distinct()

        for student in students:
            val = request.POST.get(f'attendance_{student.id}')
            if val is not None:
                try:
                    pct = float(val)
                    student.attendance = pct
                    student.save()
                    
                    # Logic for Attendance Alerts
                    cat, _ = AnnouncementCategory.objects.get_or_create(name="Condonation/Attendance")
                    if pct < 65:
                        Announcement.objects.create(
                            title="Examination Ineligibility Notice",
                            content=f"URGENT: Your overall attendance is currently {pct}%, which is below the mandatory 65%. You are currently ineligible to attend the semester examinations.",
                            category=cat,
                            priority='HIGH',
                            posted_by=request.user,
                            target_audience='INDIVIDUAL',
                            target_individual=student,
                            expiry_date=timezone.now() + timezone.timedelta(days=30)
                        )
                    elif pct < 75:
                        Announcement.objects.create(
                            title="Condonation Fee Notice",
                            content=f"Your overall attendance is currently {pct}%, which is below the required 75%. Please visit the academic office to pay the condonation fee to be eligible for exams.",
                            category=cat,
                            priority='HIGH',
                            posted_by=request.user,
                            target_audience='INDIVIDUAL',
                            target_individual=student,
                            expiry_date=timezone.now() + timezone.timedelta(days=30)
                        )
                except (ValueError, TypeError):
                    continue

        messages.success(request, f"Attendance percentages for {subject.name} have been updated directly in student profiles.")
        return redirect('faculty_attendance_subjects')

try:
    import openpyxl
except ImportError:
    openpyxl = None
from django.http import HttpResponse

class DownloadAttendanceTemplateView(LoginRequiredMixin, FacultyRequiredMixin, View):
    def get(self, request, subject_id=None):
        students = []
        subject_code = "Generic"
        
        if subject_id:
            subject = get_object_or_404(Subject, id=subject_id, faculty=request.user)
            subject_code = subject.code
            # Get students assigned to this subject (enrolled in same course and semester)
            students = User.objects.filter(
                role='STUDENT',
                enrollments__course=subject.course,
                enrollments__current_semester=subject.semester,
                enrollments__is_active=True
            ).distinct()

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Template"

        # Headers
        headers = ['STUDENT ID', 'STUDENT NAME', 'ATTENDANCE PERCENTAGE']
        ws.append(headers)

        # Data
        for student in students:
            # Populate with existing attendance percentage if available
            ws.append([student.username, student.get_full_name() or student.username, student.attendance if hasattr(student, 'attendance') else ""])

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 5)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Attendance_Template_{subject_code}.xlsx'
        wb.save(response)
        return response


class UploadAttendanceExcelView(LoginRequiredMixin, FacultyRequiredMixin, View):
    def post(self, request, subject_id=None):
        subject = None
        if subject_id:
            subject = get_object_or_404(Subject, id=subject_id, faculty=request.user)
            
        excel_file = request.FILES.get('attendance_excel')
        
        if not excel_file:
            messages.error(request, "No file uploaded.")
            return redirect('faculty_attendance_subjects')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls).")
            return redirect('faculty_attendance_subjects')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Assume first row is header
            rows = list(ws.rows)
            print("rows")
            print(rows)
            if len(rows) < 2:
                messages.error(request, "The Excel file is empty.")
                return redirect('faculty_attendance_subjects')

            # We need academic year. For simplicity, let's derive it or use a default.
            # In a real system, we might ask the user or get from settings.
            # Let's try to get it from the first student's enrollment
            academic_year = "2024-2025" # Default fallback
            
            updated_count = 0
            for row in rows[1:]: # Skip header
                student_id = row[0].value
                # student_name = row[1].value # We don't strictly need this for update
                attendance_val = row[2].value
                attendance_pct = None
                
                if student_id and attendance_val is not None:
                    try:
                        # Attempt to convert to float/int
                        attendance_pct = float(attendance_val)
                    except (ValueError, TypeError):
                        continue
                        
                    try:
                        # Map Excel 'STUDENT ID' to User 'username' field
                        clean_username = str(student_id).strip()
                        student = User.objects.get(username__iexact=clean_username, role='STUDENT')
                        
                        # Get enrollment to find academic year if possible
                        enrollment = Enrollment.objects.filter(student=student, is_active=True).first()
                        if enrollment:
                            academic_year = enrollment.academic_year
                        
                        # If subject wasn't provided in URL, try to detect it for this student
                        current_subject = subject
                        if not current_subject and enrollment:
                            # Find subject taught by this faculty that matches student's course/semester
                            current_subject = Subject.objects.filter(
                                faculty=request.user, 
                                course=enrollment.course,
                                semester=enrollment.current_semester
                            ).first()

                        if not current_subject:
                            # If we still can't find a subject for this student/faculty combo, skip
                            continue

                        # Update student attendance directly in User model
                        student.attendance = attendance_pct
                        student.save()
                        updated_count += 1
                        
                        # Logic for Attendance Alerts
                        cat, _ = AnnouncementCategory.objects.get_or_create(name="Condonation/Attendance")
                        
                        if attendance_pct < 65:
                            # Severe Alert: Ineligible for exams
                            Announcement.objects.create(
                                title="Examination Ineligibility Notice",
                                content=f"URGENT: Your overall attendance has dropped to {attendance_pct}%, which is below the mandatory 65%. You are currently ineligible to attend the semester examinations.",
                                category=cat,
                                priority='HIGH',
                                posted_by=request.user,
                                target_audience='INDIVIDUAL',
                                target_individual=student,
                                expiry_date=timezone.now() + timezone.timedelta(days=30)
                            )
                        elif attendance_pct < 75:
                            # Alert: Need Condonation
                            Announcement.objects.create(
                                title="Condonation Fee Notice",
                                content=f"Your overall attendance is currently {attendance_pct}%, which is below the required 75%. Please visit the academic office to pay the condonation fee to be eligible for exams.",
                                category=cat,
                                priority='HIGH',
                                posted_by=request.user,
                                target_audience='INDIVIDUAL',
                                target_individual=student,
                                expiry_date=timezone.now() + timezone.timedelta(days=30)
                            )
                    except User.DoesNotExist:
                        continue

            messages.success(request, f"Successfully processed {updated_count} student attendance records.")
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

        return redirect('faculty_attendance_subjects')

class DownloadDailyAttendanceTemplateView(LoginRequiredMixin, FacultyRequiredMixin, View):
    def get(self, request, subject_id=None):
        date = request.GET.get('date', timezone.now().date().isoformat())
        students = []
        subject_code = "Generic"
        
        if subject_id:
            subject = get_object_or_404(Subject, id=subject_id, faculty=request.user)
            subject_code = subject.code
            
            students = User.objects.filter(
                role='STUDENT',
                enrollments__course=subject.course,
                enrollments__current_semester=subject.semester,
                enrollments__is_active=True
            ).distinct()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance_{date}"

        ws.append(['STUDENT ID', 'STUDENT NAME', 'ATTENDANCE PERCENTAGE'])
        
        for student in students:
            ws.append([student.username, student.get_full_name() or student.username, student.attendance if hasattr(student, 'attendance') else ""])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Daily_Attendance_{subject_code}_{date}.xlsx'
        wb.save(response)
        return response

class UploadDailyAttendanceExcelView(LoginRequiredMixin, FacultyRequiredMixin, View):
    def post(self, request, subject_id=None):
        subject = None
        if subject_id:
            subject = get_object_or_404(Subject, id=subject_id, faculty=request.user)
            
        date = request.POST.get('date', timezone.now().date().isoformat())
        excel_file = request.FILES.get('daily_attendance_excel')
        
        if not excel_file:
            messages.error(request, "No file uploaded.")
            return redirect('faculty_attendance_subjects')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls).")
            if subject:
                return redirect('mark_attendance', subject_id=subject.id)
            return redirect('faculty_attendance_subjects')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.rows)
            
            if len(rows) < 2:
                messages.error(request, "The Excel file is empty.")
                if subject:
                    return redirect('mark_attendance', subject_id=subject.id)
                return redirect('faculty_attendance_subjects')
            
            count = 0
            for row in rows[1:]:
                student_id = row[0].value
                attendance_val = row[2].value
                attendance_pct = None
                
                if student_id and attendance_val is not None:
                    try:
                        # Attempt to convert to float/int
                        attendance_pct = float(attendance_val)
                    except (ValueError, TypeError):
                        continue
                        
                    try:
                        # Map Excel 'STUDENT ID' to User 'username' field
                        clean_username = str(student_id).strip()
                        student = User.objects.get(username__iexact=clean_username, role='STUDENT')
                        
                        # Update student attendance directly in User model
                        student.attendance = attendance_pct
                        student.save()
                        
                        # Alert students if performance is low
                        cat, _ = AnnouncementCategory.objects.get_or_create(name="Condonation/Attendance")
                        
                        if attendance_pct < 65:
                            # Severe Alert: Ineligible for exams
                            Announcement.objects.create(
                                title="Examination Ineligibility Notice",
                                content=f"URGENT: Your overall attendance has dropped to {attendance_pct}%, which is below the mandatory 65%. You are currently ineligible to attend the semester examinations.",
                                category=cat,
                                priority='HIGH',
                                posted_by=request.user,
                                target_audience='INDIVIDUAL',
                                target_individual=student,
                                expiry_date=timezone.now() + timezone.timedelta(days=30)
                            )
                        elif attendance_pct < 75:
                            # Alert: Need Condonation
                            Announcement.objects.create(
                                title="Condonation Fee Notice",
                                content=f"Your overall attendance is currently {attendance_pct}%, which is below the required 75%. Please visit the academic office to pay the condonation fee to be eligible for exams.",
                                category=cat,
                                priority='HIGH',
                                posted_by=request.user,
                                target_audience='INDIVIDUAL',
                                target_individual=student,
                                expiry_date=timezone.now() + timezone.timedelta(days=30)
                            )
                        
                        count += 1
                    except User.DoesNotExist:
                        continue
            
            messages.success(request, f"Successfully uploaded attendance for {count} students.")
            messages.info(request, "Note: Daily attendance upload does not update overall attendance percentage without a dedicated tracking mechanism.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

        if subject:
            return redirect(f"/attendance/faculty/mark/{subject.id}/?date={date}")
        return redirect('faculty_attendance_subjects')
